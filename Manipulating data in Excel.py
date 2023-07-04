#Created by Natanael Teixeira

import pandas as pd
import xlwings as xw
from pandas.core.reshape.merge import merge
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import formula
import xlrd 
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
import string

#Here we insert the files that will be used
df_1 = pd.read_excel('insert your file here.xlsx') #Here
df_2 = pd.read_excel('insert your file here.xlsx') #Here
df_3 = pd.read_excel('insert your file here.xlsx') #Here

#Deleting columns
deleted_columns = ['Column1', 'Column2']
df_1 = df_1.drop(columns=deleted_columns)
#inserting columns
df_1.insert(16, "Column", True)
df_1.insert(17, "Column", True)

#Saving the file
df_1.to_excel('file.xlsx', sheet_name= 'Sheet2', index=False)

#opening previously saved file
excel_file = ('file.xlsx')
book = load_workbook(excel_file)

#Renaming a column
newsheet_name = 'Sheet'
sheet = book.create_sheet(title=newsheet_name)
book.save(excel_file)

#I'm opening a second file that was necessary for my case
origin_file = openpyxl.load_workbook('file2.xlsx') #Here
origin_sheet = origin_file['Sheet']

#I am opening the first saved file again
destiny_file = openpyxl.load_workbook('file.xlsx')
destiny_sheet = destiny_file['Sheet']

#In my case, I needed to copy a database from another file to the one I created earlier
for origin_line in origin_sheet.iter_rows():
    for origin_cell in origin_line:
        destiny_cell = destiny_sheet.cell(row=origin_cell.row, column=origin_cell.column)
        destiny_cell.value = origin_cell.value
        
destiny_file.save('file1.xlsx')

#In this section, I needed him to analyze the entire copied base and compare it with the new one
df_5 = pd.read_excel('file1.xlsx', sheet_name='Sheet')
df_6 = df_2[~df_2['Id'].isin(df_5['Id'])]

#Here, I told him to create a file with the customers that were in the new base but not in mine
df_6.to_excel('file3.xlsx', index=False)

origin_file = openpyxl.load_workbook('file3.xlsx')
origin_sheet = origin_file['Sheet1']

destiny_file = openpyxl.load_workbook('file.xlsx')
destiny_sheet = destiny_file['Sheet']

#In this loop, it will take the customers that were separated in the previous file and join them with my base so that they are updated. I can't copy directly from the new base because sellers make changes that don't interest me and can complicate the calculations
last_line = destiny_sheet.max_row + 1

for i, origin_line in enumerate(origin_sheet.iter_rows(), start=1):
    if i > 1:
        line_values = [cell.value for cell in origin_line]
        destiny_sheet.append(line_values)
    
destiny_file.save('file4.xlsx')

#In this section I am asking him to copy the excel formulas from the previous base and launch them in the new base, noting that he needs to update the line in the formula as he inserts it.
workbook_origem = load_workbook('file2.xlsx') #Here
planilha_origem = workbook_origem['Sheet']


workbook_destino = load_workbook('file4.xlsx')
planilha_destino = workbook_destino['Sheet']


coluna_origem = 'Q'
coluna_destino = 'Q'
coluna_origem2 = 'R'
coluna_destino2 = 'R'


num_linhas_origem = planilha_origem.max_row
num_linhas_destino = planilha_destino.max_row


for linha in range(2, num_linhas_destino + 1):
    if linha <= num_linhas_origem:
        célula_origem = planilha_origem[f'{coluna_origem}{linha}']
        fórmula = célula_origem.value
    else:
        célula_origem = planilha_origem[f'{coluna_origem}{num_linhas_origem}']
        fórmula = célula_origem.value.replace(str(num_linhas_origem), str(linha))

    célula_destino = planilha_destino[f'{coluna_destino}{linha}']
    célula_destino.value = fórmula
    
for linha in range(2, num_linhas_destino + 1):
    if linha <= num_linhas_origem:
        célula_origem = planilha_origem[f'{coluna_origem2}{linha}']
        fórmula = célula_origem.value
    else:
        célula_origem = planilha_origem[f'{coluna_origem2}{num_linhas_origem}']
        fórmula = célula_origem.value.replace(str(num_linhas_origem), str(linha))

    célula_destino = planilha_destino[f'{coluna_destino2}{linha}']
    célula_destino.value = fórmula


workbook_destino.save('file5.xlsx')

#This is just to not stop the process.
print('Loading Styles...')

df = load_workbook('file5.xlsx')

#Here I am adjusting the font because python borders the header cells, in addition I changed other cells to bold
sheet = df['Sheet']
sheet2 = df['Sheet4']
no_border = Border()
bold_font = Font(bold=True)
header_row = sheet2[1]

for cell in sheet[1]:
    cell.border = no_border

for cell in header_row:
    cell.font = bold_font

#Here is the final file   
df.save('total.xlsx')
       
print('Successful!')